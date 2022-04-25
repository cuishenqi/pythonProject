# import requests
# import openpyxl
# import json
#
# url = 'http://hkf10.eastmoney.com/F9HKStock/GetFinanceAssetData.do?securityCode=01810.HK&comType=127000000606281483&yearList=2018,2017,2016,2015&reportTypeList=1,5,3,6&dateSearchType=1&listedType=0,1&reportTypeInScope=1&reportType=0&rotate=0&seperate=0&order=desc&cashType=1&exchangeValue=1&customSelect=0&CurrencySelect=0'
# res = requests.get(url);  # 爬取数据  //返回了json数据
# res.encoding = "utf-8"
#
# jst = json.loads(res.text)
# key = jst['resultList']
#
#
# def ExcelWrite(result):
#  mywb = openpyxl.Workbook()
#  sheet = mywb.active;  # 获取初始的sheet
#  row = 0;  # 单元格的行
#  for i in result:
#   col = 64;  # 单元格的列 从 'A' 开始
#   row += 1;
#   for j in i:
#    col += 1;
#    sheet[chr(col) + str(row)] = i[j];
#  mywb.save('C:\\Users\\Sean\\Desktop\\负债资产.xls');
#
#
# ExcelWrite(key);



# # !/usr/bin/env python
# # coding=utf-8
# import openpyxl
# from xlwt import *
#
# # 需要xlwt库的支持
# # import xlwt
# file = Workbook(encoding='utf-8')
# # 指定file以utf-8的格式打开
# # 指定打开的文件名,添加Sheet
# table = file.add_sheet('Sheet')
# # table = file.Workbook['Sheet']
#
#
# data = {
#         "1": ["张三", 150, 120, 100],
#         "2": ["李四", 90, 99, 95],
#         "3": ["王五", 60, 66, 68]
# }
# # 字典数据
#
# ldata = []
# num = [a for a in data]
# # for循环指定取出key值存入num中
# num.sort()
# # 字典数据取出后无需，需要先排序
#
# for x in num:
#  # for循环将data字典中的键和值分批的保存在ldata中
#        t = [int(x)]
#        for a in data[x]:
#            t.append(a)
#        ldata.append(t)
#
# for i, p in enumerate(ldata):
#  # 将数据写入文件,i是enumerate()函数返回的序号数
#     for j, q in enumerate(p):
#     # print i,j,q,
#         table.write(i, j, q)
# file.save('requests.xls')


import openpyxl


def write_value_open(row,col,value):
    # 先用xlrd打开源文件
    # file_path = r'D:\pythonProject\ch6\a1.requests2.1xlsx'
    # file_name = '\\requests2.1.xlsx'

    wb = openpyxl.load_workbook('requests2.1.xlsx')
    work_sheet = wb['Sheet1']
    #work_sheet['K2']='pass' #此种写法后面可能容易报错，保险起见我使用了下面的方法
    work_sheet.cell(row=row,column=col).value = value
    wb.save(file_path+file_name)

write_value_open(2,11,'pass')

